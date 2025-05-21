import openpyxl


def fill_departures_to_excel(file_path, departures_mapping):
    """
    Fill Excel with departure lane assignments.
    
    Args:
        file_path: Path to the Excel file
        departures_mapping: Dictionary mapping Lane objects to lists of bus IDs (from buses_mapped)

    Returns:
        Nothing, but modifies the Excel file in place.
    """
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Print departures_mapping
    

    # Clear column T and set header
    for row in range(1, sheet.max_row + 1):
        sheet[f"U{row}"] = None
    sheet["U1"] = "Lähtö"
    
    # Create mappings from lane to lane name
    lane_names = {}  # Dictionary to store lane:name mappings
    
    # Initialize counters outside the loop
    ulkotelit = ["U1", "U2", "U3", "U4", "U5", "U6", "U7", "U8", "U9", "U10", "U11", "U12", "U13", "U14", "U15", "U16", "U17"]
    ulkotelit_count = 0
    b_f_lanes = ["B", "C", "D", "E", "F"]
    b_f_count = 0
    h_j_lanes = ["H", "I", "J", "Yli", "Yli", "Yli", "Yli", "Yli", "Yli", "Yli", "Yli"]
    h_j_count = 0
    gkl_lanes = ["G", "K", "L", "Yli", "Yli", "Yli", "Yli", "Yli", "Yli", "Yli", "Yli"]
    gkl_count = 0

    # Build departure mapping with proper lane names
    for lane, bus_list in departures_mapping.items():
        # Determine lane name based on the bus types in the list
        first_bus = bus_list[0] if bus_list else ""
        last_bus = bus_list[-1] if bus_list else ""
        lane_size = len(bus_list)

        # Your existing if-elif conditions but now store the lane name for each bus
        if first_bus.startswith("DMS") or first_bus.startswith("DMV") and lane_size == 1:
            lane_name = "Ulkona"

        elif first_bus.startswith("STS") or first_bus.startswith("STV") and lane_size == 1:
            lane_name = ulkotelit[ulkotelit_count]
            ulkotelit_count += 1

        elif all(bus.startswith("SMS") for bus in bus_list) and lane_size == 6 or all(bus.startswith("SMV") for bus in bus_list) and lane_size == 6:
            lane_name = "A"

        elif (first_bus.startswith("DMV") or first_bus.startswith("DMS")) and (last_bus.startswith("DMV") or last_bus.startswith("DMS")) and lane_size == 6:
            lane_name = gkl_lanes[gkl_count]
            gkl_count += 1

        elif first_bus.startswith("SMS") or first_bus.startswith("SMV") and lane_size == 6:
            if last_bus.startswith("STV") or last_bus.startswith("STS"):
                lane_name = b_f_lanes[b_f_count]
                b_f_count += 1

        else:
            lane_name = h_j_lanes[h_j_count]
            h_j_count += 1

        # Store the lane name for the lane object instead of individual buses
        lane_names[lane] = lane_name

    # Pretty print the lane assignments
    print("\nLane Assignments:")
    print("----------------")
    
    # Group lanes by their name prefix
    main_lanes = {k:v for k,v in lane_names.items() if v in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']}
    outside_lanes = {k:v for k,v in lane_names.items() if v.startswith('U')}
    other_lanes = {k:v for k,v in lane_names.items() if v not in main_lanes.values() and v not in outside_lanes.values()}

    # Print main lanes
    print("\nMain Lanes:")
    for lane, name in sorted(main_lanes.items(), key=lambda x: x[1]):
        first_bus = departures_mapping[lane][0]
        last_bus = departures_mapping[lane][-1]
        print(f"Lane {name}: {first_bus} -> {last_bus}")

    # Print outside charging spots
    print("\nOutside Charging Spots:")
    for lane, name in sorted(outside_lanes.items(), key=lambda x: x[1]):
        bus = departures_mapping[lane][0]  # Only one bus for outside spots
        print(f"Spot {name}: {bus}")

    # Print other assignments
    if other_lanes:
        print("\nOther Assignments:")
        for lane, name in sorted(other_lanes.items(), key=lambda x: x[1]):
            buses = departures_mapping[lane]
            print(f"{name}: {', '.join(buses)}")

    # Track last occurrences of each bus in Excel
    bus_occurrences = {}
    for row in range(2, sheet.max_row + 1):
        bus_id = sheet[f"B{row}"].value
        if bus_id:
            bus_occurrences[bus_id] = row

    # Create bus to lane name mapping
    bus_to_lane = {}
    for lane, bus_list in departures_mapping.items():
        lane_name = lane_names[lane]
        for i, bus in enumerate(bus_list):
            # For main lanes (A-L), add position number
            if lane_name in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                bus_to_lane[bus] = f"{lane_name}{i+1}"
            else:
                bus_to_lane[bus] = lane_name

    # Track written buses and find missing ones
    written_buses = set()
    all_buses = set()
    
    # Collect all buses from departures_mapping
    for lane, bus_list in departures_mapping.items():
        all_buses.update(bus_list)

    # Write lane assignments to Excel and track written buses
    for bus_id, last_row in bus_occurrences.items():
        if bus_id in bus_to_lane:
            sheet[f"U{last_row}"] = bus_to_lane[bus_id]
            written_buses.add(bus_id)

    # Find missing buses
    missing_buses = all_buses - written_buses
    if missing_buses:
        print("\nWARNING: The following buses were not written to Excel:")
        for bus in sorted(missing_buses):
            print(f"- {bus} (Lane: {bus_to_lane.get(bus, 'Unknown')})")
            # Add these buses to the end of Excel
            next_row = sheet.max_row + 1
            sheet[f"B{next_row}"] = bus
            sheet[f"U{next_row}"] = bus_to_lane[bus]
            print(f"  -> Added to row {next_row}")

    workbook.save(file_path)
    
    # Print summary
    print(f"\nSummary:")
    print(f"Total buses: {len(all_buses)}")
    print(f"Written buses: {len(written_buses)}")
    print(f"Added missing buses: {len(missing_buses)}")


def fill_arrivals_to_excel(file_path, arrivals_mapping):
    """
        Fill Excel with arrival lane assignments.
        
        Args:
            file_path: Path to the Excel file
            
            arrivals_mapping: Dictionary mapping Lane objects to lists of bus IDs (from lanes)

        Returns:
            Nothing, but modifies the Excel file in place.
    """
   
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Print departures_mapping
    

    # Clear column T and set header
    for row in range(1, sheet.max_row + 1):
        sheet[f"T{row}"] = None
    sheet["T1"] = "Saapuminen"
    
    # Create mappings from lane to lane name
    lane_names = {}  # Dictionary to store lane:name mappings
    
    # Initialize counters outside the loop
    ulkotelit = ["U1", "U2", "U3", "U4", "U5", "U6", "U7", "U8", "U9", "U10", "U11", "U12", "U13", "U14", "U15", "U16", "U17"]
    ulkotelit_count = 0
    b_f_lanes = ["B", "C", "D", "E", "F"]
    b_f_count = 0
    h_j_lanes = ["H", "I", "J", "Yli", "Yli", "Yli", "Yli", "Yli", "Yli", "Yli", "Yli"]
    h_j_count = 0
    gkl_lanes = ["G", "K", "L", "Yli", "Yli", "Yli", "Yli", "Yli", "Yli", "Yli", "Yli"]
    gkl_count = 0

    # Build departure mapping with proper lane names
    for lane, bus_list in arrivals_mapping.items():
        # Determine lane name based on the bus types in the list
        first_bus = bus_list[0] if bus_list else ""
        last_bus = bus_list[-1] if bus_list else ""
        lane_size = len(bus_list)

        # Your existing if-elif conditions but now store the lane name for each bus
        if first_bus.startswith("DMS") or first_bus.startswith("DMV") and lane_size == 1:
            lane_name = "Ulkona"

        elif first_bus.startswith("STS") or first_bus.startswith("STV") and lane_size == 1:
            lane_name = ulkotelit[ulkotelit_count]
            ulkotelit_count += 1

        elif all(bus.startswith("SMS") for bus in bus_list) and lane_size == 6 or all(bus.startswith("SMV") for bus in bus_list) and lane_size == 6:
            lane_name = "A"

        elif (first_bus.startswith("DMV") or first_bus.startswith("DMS")) and (last_bus.startswith("DMV") or last_bus.startswith("DMS")) and lane_size == 6:
            lane_name = gkl_lanes[gkl_count]
            gkl_count += 1

        elif first_bus.startswith("SMS") or first_bus.startswith("SMV") and lane_size == 6:
            if last_bus.startswith("STV") or last_bus.startswith("STS"):
                lane_name = b_f_lanes[b_f_count]
                b_f_count += 1

        else:
            lane_name = h_j_lanes[h_j_count]
            h_j_count += 1

        # Store the lane name for the lane object instead of individual buses
        lane_names[lane] = lane_name

    # Pretty print the lane assignments
    print("\nLane Assignments:")
    print("----------------")
    
    # Group lanes by their name prefix
    main_lanes = {k:v for k,v in lane_names.items() if v in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']}
    outside_lanes = {k:v for k,v in lane_names.items() if v.startswith('U')}
    other_lanes = {k:v for k,v in lane_names.items() if v not in main_lanes.values() and v not in outside_lanes.values()}

    # Print main lanes
    print("\nMain Lanes:")
    for lane, name in sorted(main_lanes.items(), key=lambda x: x[1]):
        first_bus = arrivals_mapping[lane][0]
        last_bus = arrivals_mapping[lane][-1]
        print(f"Lane {name}: {first_bus} -> {last_bus}")

    # Print outside charging spots
    print("\nOutside Charging Spots:")
    for lane, name in sorted(outside_lanes.items(), key=lambda x: x[1]):
        bus = arrivals_mapping[lane][0]  # Only one bus for outside spots
        print(f"Spot {name}: {bus}")

    # Print other assignments
    if other_lanes:
        print("\nOther Assignments:")
        for lane, name in sorted(other_lanes.items(), key=lambda x: x[1]):
            buses = arrivals_mapping[lane]
            print(f"{name}: {', '.join(buses)}")

    # Track first occurrences of each bus in Excel (instead of last)
    bus_occurrences = {}
    for row in range(2, sheet.max_row + 1):
        bus_id = sheet[f"B{row}"].value
        if bus_id and bus_id not in bus_occurrences:  # Only take first occurrence
            bus_occurrences[bus_id] = row

    # Create bus to lane name mapping
    bus_to_lane = {}
    for lane, bus_list in arrivals_mapping.items():
        lane_name = lane_names[lane]
        for i, bus in enumerate(bus_list):
            # For main lanes (A-L), add position number
            if lane_name in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                bus_to_lane[bus] = f"{lane_name}{i+1}"
            else:
                bus_to_lane[bus] = lane_name

    # Track written buses and find missing ones
    written_buses = set()
    all_buses = set()
    
    # Collect all buses from departures_mapping
    for lane, bus_list in arrivals_mapping.items():
        all_buses.update(bus_list)

    # Write lane assignments to Excel and track written buses
    for bus_id, first_row in bus_occurrences.items():
        if bus_id in bus_to_lane:
            sheet[f"T{first_row}"] = bus_to_lane[bus_id]
            written_buses.add(bus_id)

    # Find missing buses
    missing_buses = all_buses - written_buses
    if missing_buses:
        print("\nWARNING: The following buses were not written to Excel:")
        for bus in sorted(missing_buses):
            print(f"- {bus} (Lane: {bus_to_lane.get(bus, 'Unknown')})")
            # Add these buses to the end of Excel
            next_row = sheet.max_row + 1
            sheet[f"B{next_row}"] = bus
            sheet[f"U{next_row}"] = bus_to_lane[bus]
            print(f"  -> Added to row {next_row}")

    workbook.save(file_path)
    
    # Print summary
    print(f"\nSummary:")
    print(f"Total buses: {len(all_buses)}")
    print(f"Written buses: {len(written_buses)}")
    print(f"Added missing buses: {len(missing_buses)}")